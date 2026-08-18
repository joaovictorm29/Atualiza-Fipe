"""Atualiza a coluna VALOR FIPE CHASSI com segurança.

O programa nunca escreve um preço quando há mais de um modelo/ano FIPE
compatível. Nesses casos, registra os candidatos e aguarda confirmação humana.
Depois que CÓDIGO FIPE e ANO FIPE estiverem preenchidos uma vez, as execuções
seguintes consultam a FIPE diretamente pelos códigos, sem comparação textual.
"""

from __future__ import annotations

import csv
import os
import re
import time
import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

import openpyxl
import requests


CAMINHO_PLANILHA = Path(r"C:\Users\user\Downloads\autualizaFipe\Planilha Veículos c- depreciação.xlsx")
CAMINHO_SAIDA = Path(r"C:\Users\user\Downloads\Planilha Veículos c- depreciação - atualizada.xlsx")
CAMINHO_PENDENCIAS = Path(r"C:\Users\user\Downloads\pendencias_fipe.csv")
ABA_VEICULOS = "VEICULOS"
LINHA_CABECALHO = 2
LINHA_INICIO_DADOS = 3
API_BASE = "https://fipe.parallelum.com.br/api/v2"
TIMEOUT_SEGUNDOS = 30
TENTATIVAS_API = 3

# São acrescentadas automaticamente à direita da tabela, se ainda não existirem.
COLUNAS_CONTROLE = [
    "MODELO CODE FIPE",
    "CÓDIGO FIPE",
    "ANO FIPE",
    "MODELO FIPE",
    "STATUS FIPE",
    "MOTIVO FIPE",
    "REFERÊNCIA FIPE",
    "CANDIDATOS FIPE",
]

# Apelidos controlados. Inclua aqui somente equivalências conhecidas, nunca
# aproximações livres de marca.
ALIASES_MARCA = {
    "VW": "VOLKSWAGEN",
    "VW VOLKSWAGEN": "VOLKSWAGEN",
}

PALAVRAS_NEUTRAS = {
    "A", "BASCULANTE", "CABINE", "CAMINHAO", "CARGA", "COM",
    "COMPLEMENTAR", "DE", "E", "IMPLEMENTO", "LITROS", "METROS",
    "MUNCK", "PARA", "PIPA", "TUDO", "VEICULO", "VOLKSWAGEN", "VW",
}


@dataclass(frozen=True)
class Veiculo:
    linha: int
    codigo_ativo: str
    tipo: str
    modelo_crlv: str
    marca: str
    categoria: str
    ano: Any


class FipeErro(RuntimeError):
    pass


def normalizar(texto: Any) -> str:
    texto = "" if texto is None else str(texto)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto.upper().strip()


def chave_marca(texto: Any) -> str:
    chave = normalizar(texto).replace("-", " ").replace("/", " ")
    chave = re.sub(r"\bVW\b", "", chave)
    chave = re.sub(r"\s+", " ", chave).strip()
    return ALIASES_MARCA.get(chave, chave)


def normalizar_modelo(texto: Any) -> str:
    """Remove placa e uniformiza grafias que não alteram o modelo."""
    texto = normalizar(texto)
    texto = re.sub(r"\([A-Z]{3}-[A-Z0-9]{4}\)", " ", texto)
    texto = re.sub(r"\bCONSTEL\.?\b", "CONSTELLATION", texto)
    texto = re.sub(r"\bROBUSTER\b", "ROBUST", texto)
    texto = re.sub(r"\bCDSRV\b", "CD SRV", texto)
    # Deixa códigos de caminhão comparáveis: 31.320 e 31-320 viram 31-320.
    texto = re.sub(r"\b(\d{1,2})[.]\s*(\d{3})\b", r"\1-\2", texto)
    return re.sub(r"\s+", " ", texto).strip()


def tokens(texto: str) -> set[str]:
    return set(re.findall(r"[A-Z0-9]+(?:[.-][A-Z0-9]+)?", texto))


def codigos_tecnicos(texto: str) -> set[str]:
    return {f"{a}-{b}" for a, b in re.findall(r"\b(\d{1,2})[.-](\d{3})\b", texto)}


def tracoes(texto: str) -> set[str]:
    return set(re.findall(r"\b[468]X[246]\b", texto))


def termos_relevantes(texto: str) -> set[str]:
    resultado = set()
    for termo in tokens(texto):
        if termo in PALAVRAS_NEUTRAS or re.fullmatch(r"\d+(?:[.-]\d+)?", termo):
            continue
        if re.fullmatch(r"\d+P", termo):
            continue
        resultado.add(termo)
    return resultado


def ano_desejado(valor: Any) -> int | None:
    """Em 2025/2026, usa 2026: o ano/modelo que a FIPE normalmente lista."""
    encontrados = re.findall(r"(?:19|20)\d{2}", str(valor))
    return int(encontrados[-1]) if encontrados else None


def preco_para_numero(preco: str) -> float:
    return float(preco.replace("R$", "").replace(".", "").replace(",", ".").strip())


class ClienteFipe:
    def __init__(self) -> None:
        self.session = requests.Session()
        token = os.getenv("FIPE_API_TOKEN")
        if token:
            self.session.headers["X-Subscription-Token"] = token

    def get(self, caminho: str) -> Any:
        ultimo_erro: Exception | None = None
        for tentativa in range(TENTATIVAS_API):
            try:
                resposta = self.session.get(
                    f"{API_BASE}{caminho}", timeout=TIMEOUT_SEGUNDOS
                )
                resposta.raise_for_status()
                return resposta.json()
            except requests.RequestException as erro:
                ultimo_erro = erro
                if tentativa + 1 < TENTATIVAS_API:
                    time.sleep(2 ** tentativa)
        raise FipeErro(str(ultimo_erro))


cliente = ClienteFipe()


@lru_cache(maxsize=None)
def marcas(categoria: str) -> tuple[tuple[str, str], ...]:
    resposta = cliente.get(f"/{categoria}/brands")
    return tuple((str(item["code"]), item["name"]) for item in resposta)


@lru_cache(maxsize=None)
def modelos(categoria: str, marca_code: str) -> tuple[tuple[str, str], ...]:
    resposta = cliente.get(f"/{categoria}/brands/{marca_code}/models")
    return tuple((str(item["code"]), item["name"]) for item in resposta)


@lru_cache(maxsize=None)
def anos_do_modelo(
    categoria: str, marca_code: str, modelo_code: str
) -> tuple[tuple[str, str], ...]:
    resposta = cliente.get(
        f"/{categoria}/brands/{marca_code}/models/{modelo_code}/years"
    )
    return tuple((str(item["code"]), item["name"]) for item in resposta)


@lru_cache(maxsize=None)
def detalhe_por_modelo(
    categoria: str, marca_code: str, modelo_code: str, ano_code: str
) -> dict[str, Any]:
    return cliente.get(
        f"/{categoria}/brands/{marca_code}/models/{modelo_code}/years/{ano_code}"
    )


@lru_cache(maxsize=None)
def anos_por_codigo_fipe(categoria: str, codigo_fipe: str) -> tuple[tuple[str, str], ...]:
    resposta = cliente.get(f"/{categoria}/{codigo_fipe}/years")
    return tuple((str(item["code"]), item["name"]) for item in resposta)


@lru_cache(maxsize=None)
def detalhe_por_codigo_fipe(
    categoria: str, codigo_fipe: str, ano_code: str
) -> dict[str, Any]:
    return cliente.get(f"/{categoria}/{codigo_fipe}/years/{ano_code}")


def encontrar_marca_exata(categoria: str, texto_marca: str) -> tuple[str, str] | None:
    alvo = chave_marca(texto_marca)
    encontrados = [marca for marca in marcas(categoria) if chave_marca(marca[1]) == alvo]
    return encontrados[0] if len(encontrados) == 1 else None


def pontuar_modelos(veiculo: Veiculo, modelos_fipe: tuple[tuple[str, str], ...]) -> list[tuple[float, str, str]]:
    """Retorna somente modelos que não contradizem o CRLV.

    Código técnico e tração são filtros eliminatórios. Similaridade textual é
    usada apenas para ordenar os candidatos restantes, nunca para "adivinhar"
    um modelo incompatível.
    """
    texto_origem = normalizar_modelo(f"{veiculo.tipo} {veiculo.modelo_crlv}")
    codigos_origem = codigos_tecnicos(texto_origem)
    tracoes_origem = tracoes(texto_origem)
    termos_origem = termos_relevantes(texto_origem)
    pontuados: list[tuple[float, str, str]] = []

    for modelo_code, nome_modelo in modelos_fipe:
        texto_fipe = normalizar_modelo(nome_modelo)
        codigos_fipe = codigos_tecnicos(texto_fipe)
        tracoes_fipe = tracoes(texto_fipe)

        if codigos_origem and not (codigos_origem & codigos_fipe):
            continue
        if tracoes_origem and (not tracoes_fipe or not (tracoes_origem & tracoes_fipe)):
            continue

        termos_fipe = termos_relevantes(texto_fipe)
        comuns = termos_origem & termos_fipe
        if not codigos_origem and len(comuns) < 2:
            continue

        # Pontos explicitamente explicáveis no relatório de revisão.
        pontuacao = len(comuns) * 18
        pontuacao += len(codigos_origem & codigos_fipe) * 80
        pontuacao += len(tracoes_origem & tracoes_fipe) * 30

        # Favorece o texto mais parecido depois de todos os filtros objetivos.
        pontuacao += 20 * __import__("difflib").SequenceMatcher(
            None, texto_origem, texto_fipe
        ).ratio()
        pontuados.append((pontuacao, modelo_code, nome_modelo))

    return sorted(pontuados, key=lambda item: item[0], reverse=True)


def anos_compativeis(
    categoria: str, marca_code: str, modelo_code: str, ano: int
) -> list[tuple[str, str]]:
    return [
        item for item in anos_do_modelo(categoria, marca_code, modelo_code)
        if ano_desejado(item[1]) == ano
    ]


def resolver_automaticamente(
    veiculo: Veiculo, marca_code: str
) -> tuple[dict[str, Any] | None, list[str], str]:
    ano = ano_desejado(veiculo.ano)
    if ano is None:
        return None, [], "ANO inválido; informe o ano/modelo com quatro dígitos."

    candidatos_modelo = pontuar_modelos(
        veiculo, modelos(veiculo.categoria, marca_code)
    )
    if not candidatos_modelo:
        return None, [], "Nenhum modelo FIPE compatível com os dados do CRLV."

    # Só consulta os melhores candidatos. Empates reais seguem para revisão,
    # em vez de aumentar desnecessariamente as chamadas à API.
    melhor_pontuacao = candidatos_modelo[0][0]
    melhores = [item for item in candidatos_modelo if item[0] >= melhor_pontuacao - 18][:8]

    opcoes: list[tuple[str, str, str, str]] = []
    for _, modelo_code, nome_modelo in melhores:
        for ano_code, nome_ano in anos_compativeis(
            veiculo.categoria, marca_code, modelo_code, ano
        ):
            opcoes.append((modelo_code, nome_modelo, ano_code, nome_ano))

    texto_candidatos = [
        f"{modelo_code} | {nome_modelo} | {nome_ano} ({ano_code})"
        for modelo_code, nome_modelo, ano_code, nome_ano in opcoes
    ]
    if len(opcoes) != 1:
        motivo = (
            "Mais de uma combinação modelo/ano compatível; confirme o código FIPE."
            if opcoes else "Nenhuma versão desse modelo está disponível para o ano informado."
        )
        return None, texto_candidatos, motivo

    modelo_code, _, ano_code, _ = opcoes[0]
    detalhe = {
        **detalhe_por_modelo(veiculo.categoria, marca_code, modelo_code, ano_code),
        "_modelo_code": modelo_code,
        "_ano_code": ano_code,
    }
    return detalhe, texto_candidatos, "Correspondência única por marca, modelo e ano."


def confirmar_por_codigo(
    veiculo: Veiculo, codigo_fipe: str, ano_fipe: str
) -> tuple[dict[str, Any] | None, str]:
    ano = ano_desejado(veiculo.ano)
    anos = anos_por_codigo_fipe(veiculo.categoria, codigo_fipe)
    if ano_fipe:
        encontrados = [item for item in anos if item[0] == ano_fipe]
    else:
        encontrados = [item for item in anos if ano_desejado(item[1]) == ano]

    if len(encontrados) != 1:
        return None, "CÓDIGO FIPE informado, mas ANO FIPE não é único/compatível."
    ano_code = encontrados[0][0]
    detalhe = {
        **detalhe_por_codigo_fipe(veiculo.categoria, codigo_fipe, ano_code),
        "_ano_code": ano_code,
    }
    return detalhe, "Atualizado pelo código FIPE confirmado."


def confirmar_por_modelo(
    veiculo: Veiculo, marca_code: str, modelo_code: str, ano_fipe: str
) -> tuple[dict[str, Any] | None, str]:
    """Usa a seleção humana de MODELO CODE FIPE + ANO FIPE com segurança."""
    ano = ano_desejado(veiculo.ano)
    anos = anos_do_modelo(veiculo.categoria, marca_code, modelo_code)
    if ano_fipe:
        encontrados = [item for item in anos if item[0] == ano_fipe]
    else:
        encontrados = [item for item in anos if ano_desejado(item[1]) == ano]

    if len(encontrados) != 1:
        return None, "MODELO CODE FIPE informado, mas ANO FIPE não é único/compatível."

    ano_code = encontrados[0][0]
    detalhe = {
        **detalhe_por_modelo(veiculo.categoria, marca_code, modelo_code, ano_code),
        "_modelo_code": modelo_code,
        "_ano_code": ano_code,
    }
    return detalhe, "Atualizado pela combinação modelo/ano confirmada."


def cabecalhos(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        normalizar(celula.value): celula.column
        for celula in ws[LINHA_CABECALHO]
        if celula.value is not None
    }


def garantir_colunas_controle(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    existentes = cabecalhos(ws)
    ultima_coluna = max(existentes.values())
    for nome in COLUNAS_CONTROLE:
        chave = normalizar(nome)
        if chave not in existentes:
            ultima_coluna += 1
            ws.cell(LINHA_CABECALHO, ultima_coluna).value = nome
            existentes[chave] = ultima_coluna
    return existentes


def escrever(ws, linha: int, colunas: dict[str, int], nome: str, valor: Any) -> None:
    ws.cell(linha, colunas[normalizar(nome)]).value = valor


def valor(ws, linha: int, colunas: dict[str, int], nome: str) -> Any:
    return ws.cell(linha, colunas[normalizar(nome)]).value


def executar() -> None:
    if not CAMINHO_PLANILHA.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {CAMINHO_PLANILHA.resolve()}")

    wb = openpyxl.load_workbook(CAMINHO_PLANILHA)
    ws = wb[ABA_VEICULOS]
    colunas = garantir_colunas_controle(ws)
    obrigatorias = ["CÓD ATIVO", "TIPO", "MODELO", "MARCA", "CATEGORIA FIPE", "ANO", "VALOR FIPE CHASSI"]
    ausentes = [nome for nome in obrigatorias if normalizar(nome) not in colunas]
    if ausentes:
        raise KeyError(f"Cabeçalhos obrigatórios ausentes: {', '.join(ausentes)}")

    pendencias: list[dict[str, Any]] = []
    atualizados = 0
    revisao = 0

    for linha in range(LINHA_INICIO_DADOS, ws.max_row + 1):
        codigo_ativo = valor(ws, linha, colunas, "CÓD ATIVO")
        if codigo_ativo is None or "TOTAL" in normalizar(codigo_ativo):
            continue

        categoria = normalizar(valor(ws, linha, colunas, "CATEGORIA FIPE")).lower()
        veiculo = Veiculo(
            linha=linha,
            codigo_ativo=str(codigo_ativo),
            tipo=str(valor(ws, linha, colunas, "TIPO") or ""),
            modelo_crlv=str(valor(ws, linha, colunas, "MODELO") or ""),
            marca=str(valor(ws, linha, colunas, "MARCA") or ""),
            categoria=categoria,
            ano=valor(ws, linha, colunas, "ANO"),
        )

        if categoria not in {"cars", "trucks"}:
            escrever(ws, linha, colunas, "STATUS FIPE", "REVISÃO NECESSÁRIA")
            escrever(ws, linha, colunas, "MOTIVO FIPE", "CATEGORIA FIPE deve ser cars ou trucks.")
            revisao += 1
            continue
        if not all([veiculo.marca, veiculo.modelo_crlv, veiculo.ano]):
            escrever(ws, linha, colunas, "STATUS FIPE", "DADO INCOMPLETO")
            escrever(ws, linha, colunas, "MOTIVO FIPE", "Preencha MARCA, MODELO e ANO.")
            revisao += 1
            continue

        try:
            codigo_fipe_confirmado = str(valor(ws, linha, colunas, "CÓDIGO FIPE") or "").strip()
            modelo_code_confirmado = str(valor(ws, linha, colunas, "MODELO CODE FIPE") or "").strip()
            ano_fipe_confirmado = str(valor(ws, linha, colunas, "ANO FIPE") or "").strip()

            if codigo_fipe_confirmado:
                detalhe, motivo = confirmar_por_codigo(
                    veiculo, codigo_fipe_confirmado, ano_fipe_confirmado
                )
                candidatos = []
            else:
                marca = encontrar_marca_exata(categoria, veiculo.marca)
                if marca is None:
                    detalhe, candidatos, motivo = None, [], "Marca não encontrada de forma exata; cadastre um apelido conhecido."
                elif modelo_code_confirmado:
                    detalhe, motivo = confirmar_por_modelo(
                        veiculo, marca[0], modelo_code_confirmado, ano_fipe_confirmado
                    )
                    candidatos = []
                else:
                    detalhe, candidatos, motivo = resolver_automaticamente(veiculo, marca[0])

            if detalhe is None:
                escrever(ws, linha, colunas, "STATUS FIPE", "REVISÃO NECESSÁRIA")
                escrever(ws, linha, colunas, "MOTIVO FIPE", motivo)
                escrever(ws, linha, colunas, "CANDIDATOS FIPE", "\n".join(candidatos))
                pendencias.append({
                    "linha": linha,
                    "codigo_ativo": veiculo.codigo_ativo,
                    "marca": veiculo.marca,
                    "modelo_crlv": veiculo.modelo_crlv,
                    "ano": veiculo.ano,
                    "motivo": motivo,
                    "candidatos": " | ".join(candidatos),
                })
                revisao += 1
                print(f"[REVISAR] {veiculo.codigo_ativo}: {motivo}")
                continue

            escrever(ws, linha, colunas, "VALOR FIPE CHASSI", preco_para_numero(detalhe["price"]))
            ws.cell(linha, colunas[normalizar("VALOR FIPE CHASSI")]).number_format = 'R$ #,##0.00'
            escrever(ws, linha, colunas, "CÓDIGO FIPE", detalhe["codeFipe"])
            if detalhe.get("_modelo_code"):
                escrever(ws, linha, colunas, "MODELO CODE FIPE", detalhe["_modelo_code"])
            escrever(ws, linha, colunas, "ANO FIPE", detalhe["_ano_code"])
            escrever(ws, linha, colunas, "MODELO FIPE", detalhe["model"])
            escrever(ws, linha, colunas, "REFERÊNCIA FIPE", detalhe["referenceMonth"])
            escrever(ws, linha, colunas, "CANDIDATOS FIPE", "")
            escrever(ws, linha, colunas, "STATUS FIPE", "ATUALIZADO")
            escrever(ws, linha, colunas, "MOTIVO FIPE", motivo)
            atualizados += 1
            print(f"[OK] {veiculo.codigo_ativo}: {detalhe['model']} -> {detalhe['price']}")

        except FipeErro as erro:
            escrever(ws, linha, colunas, "STATUS FIPE", "ERRO DE API")
            escrever(ws, linha, colunas, "MOTIVO FIPE", str(erro))
            revisao += 1
            print(f"[ERRO] {veiculo.codigo_ativo}: {erro}")

    # Faz as fórmulas que dependem do valor FIPE serem recalculadas ao abrir no Excel.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        # Compatibilidade com versões antigas do openpyxl.
        pass
    wb.save(CAMINHO_SAIDA)

    with CAMINHO_PENDENCIAS.open("w", newline="", encoding="utf-8-sig") as arquivo:
        campos = ["linha", "codigo_ativo", "marca", "modelo_crlv", "ano", "motivo", "candidatos"]
        escritor = csv.DictWriter(arquivo, fieldnames=campos, delimiter=";")
        escritor.writeheader()
        escritor.writerows(pendencias)

    print(f"\nConcluído: {atualizados} atualizado(s), {revisao} para revisão.")
    print(f"Planilha: {CAMINHO_SAIDA.resolve()}")
    print(f"Pendências: {CAMINHO_PENDENCIAS.resolve()}")


if __name__ == "__main__":
    executar()
