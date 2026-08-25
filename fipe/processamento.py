from typing import Any

import openpyxl

from relatorios.pendencias import salvar_pendencias

from .cliente_api import FipeErro
from .config import (
    ABA_VEICULOS,
    CAMINHO_PENDENCIAS,
    CAMINHO_PLANILHA,
    LINHA_INICIO_DADOS,
)
from .models import Veiculo
from .normalizacao import normalizar, preco_para_numero
from .planilha import garantir_colunas_controle, escrever, valor
from .resolver import (
    confirmar_por_codigo,
    confirmar_por_modelo,
    encontrar_marca_exata,
    resolver_automaticamente,
)


def executar() -> None:
    if not CAMINHO_PLANILHA.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {CAMINHO_PLANILHA.resolve()}")

    wb = openpyxl.load_workbook(CAMINHO_PLANILHA)
    ws = wb[ABA_VEICULOS]
    colunas = garantir_colunas_controle(ws)
    obrigatorias = ["CÓD ATIVO", "TIPO", "MODELO", "MARCA", "CATEGORIA FIPE", "ANO", "VALOR FIPE CHASSI"]
    ausentes = [nome for nome in obrigatorias if normalizar(nome) not in colunas]
    if ausentes:
        raise KeyError(f"Cabeçalhos obrigatórios ausentes: {', '.join(ausentes)}")

    pendencias: list[dict[str, Any]] = []
    atualizados = 0
    revisao = 0

    for linha in range(LINHA_INICIO_DADOS, ws.max_row + 1):
        codigo_ativo = valor(ws, linha, colunas, "CÓD ATIVO")
        if codigo_ativo is None or "TOTAL" in normalizar(codigo_ativo):
            continue

        categoria = normalizar(valor(ws, linha, colunas, "CATEGORIA FIPE")).lower()
        veiculo = Veiculo(
            linha=linha,
            codigo_ativo=str(codigo_ativo),
            tipo=str(valor(ws, linha, colunas, "TIPO") or ""),
            modelo_crlv=str(valor(ws, linha, colunas, "MODELO") or ""),
            marca=str(valor(ws, linha, colunas, "MARCA") or ""),
            categoria=categoria,
            ano=valor(ws, linha, colunas, "ANO"),
        )

        if categoria not in {"cars", "trucks"}:
            escrever(ws, linha, colunas, "STATUS FIPE", "REVISÃO NECESSÁRIA")
            escrever(ws, linha, colunas, "MOTIVO FIPE", "CATEGORIA FIPE deve ser cars ou trucks.")
            revisao += 1
            continue
        if not all([veiculo.marca, veiculo.modelo_crlv, veiculo.ano]):
            escrever(ws, linha, colunas, "STATUS FIPE", "DADO INCOMPLETO")
            escrever(ws, linha, colunas, "MOTIVO FIPE", "Preencha MARCA, MODELO e ANO.")
            revisao += 1
            continue

        try:
            codigo_fipe_confirmado = str(valor(ws, linha, colunas, "CÓDIGO FIPE") or "").strip()
            modelo_code_confirmado = str(valor(ws, linha, colunas, "MODELO CODE FIPE") or "").strip()
            ano_fipe_confirmado = str(valor(ws, linha, colunas, "ANO FIPE") or "").strip()

            if codigo_fipe_confirmado:
                detalhe, motivo = confirmar_por_codigo(
                    veiculo, codigo_fipe_confirmado, ano_fipe_confirmado
                )
                candidatos = []
            else:
                marca = encontrar_marca_exata(categoria, veiculo.marca)
                if marca is None:
                    detalhe, candidatos, motivo = None, [], "Marca não encontrada de forma exata; cadastre um apelido conhecido."
                elif modelo_code_confirmado:
                    detalhe, motivo = confirmar_por_modelo(
                        veiculo, marca[0], modelo_code_confirmado, ano_fipe_confirmado
                    )
                    candidatos = []
                else:
                    detalhe, candidatos, motivo = resolver_automaticamente(veiculo, marca[0])

            if detalhe is None:
                escrever(ws, linha, colunas, "STATUS FIPE", "REVISÃO NECESSÁRIA")
                escrever(ws, linha, colunas, "MOTIVO FIPE", motivo)
                escrever(ws, linha, colunas, "CANDIDATOS FIPE", "\n".join(candidatos))
                pendencias.append({
                    "linha": linha,
                    "codigo_ativo": veiculo.codigo_ativo,
                    "marca": veiculo.marca,
                    "modelo_crlv": veiculo.modelo_crlv,
                    "ano": veiculo.ano,
                    "motivo": motivo,
                    "candidatos": " | ".join(candidatos),
                })
                revisao += 1
                print(f"[REVISAR] {veiculo.codigo_ativo}: {motivo}")
                continue

            escrever(ws, linha, colunas, "VALOR FIPE CHASSI", preco_para_numero(detalhe["price"]))
            ws.cell(linha, colunas[normalizar("VALOR FIPE CHASSI")]).number_format = 'R$ #,##0.00'
            escrever(ws, linha, colunas, "CÓDIGO FIPE", detalhe["codeFipe"])
            if detalhe.get("_modelo_code"):
                escrever(ws, linha, colunas, "MODELO CODE FIPE", detalhe["_modelo_code"])
            escrever(ws, linha, colunas, "ANO FIPE", detalhe["_ano_code"])
            escrever(ws, linha, colunas, "MODELO FIPE", detalhe["model"])
            escrever(ws, linha, colunas, "REFERÊNCIA FIPE", detalhe["referenceMonth"])
            escrever(ws, linha, colunas, "CANDIDATOS FIPE", "")
            escrever(ws, linha, colunas, "STATUS FIPE", "ATUALIZADO")
            escrever(ws, linha, colunas, "MOTIVO FIPE", motivo)
            atualizados += 1
            print(f"[OK] {veiculo.codigo_ativo}: {detalhe['model']} -> {detalhe['price']}")

        except FipeErro as erro:
            escrever(ws, linha, colunas, "STATUS FIPE", "ERRO DE API")
            escrever(ws, linha, colunas, "MOTIVO FIPE", str(erro))
            revisao += 1
            print(f"[ERRO] {veiculo.codigo_ativo}: {erro}")

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        pass
    wb.save(CAMINHO_PLANILHA)

    salvar_pendencias(CAMINHO_PENDENCIAS, pendencias)

    print(f"\nConcluído: {atualizados} atualizado(s), {revisao} para revisão.")
    print(f"Planilha atualizada: {CAMINHO_PLANILHA.resolve()}")
    print(f"Pendências: {CAMINHO_PENDENCIAS.resolve()}")
