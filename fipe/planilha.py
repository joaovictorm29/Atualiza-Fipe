from typing import Any

import openpyxl

from .config import COLUNAS_CONTROLE, LINHA_CABECALHO
from .normalizacao import normalizar


def cabecalhos(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        normalizar(celula.value): celula.column
        for celula in ws[LINHA_CABECALHO]
        if celula.value is not None
    }


def garantir_colunas_controle(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    existentes = cabecalhos(ws)
    ultima_coluna = max(existentes.values())
    for nome in COLUNAS_CONTROLE:
        chave = normalizar(nome)
        if chave not in existentes:
            ultima_coluna += 1
            ws.cell(LINHA_CABECALHO, ultima_coluna).value = nome
            existentes[chave] = ultima_coluna
    return existentes


def escrever(ws, linha: int, colunas: dict[str, int], nome: str, valor: Any) -> None:
    ws.cell(linha, colunas[normalizar(nome)]).value = valor


def valor(ws, linha: int, colunas: dict[str, int], nome: str) -> Any:
    return ws.cell(linha, colunas[normalizar(nome)]).value
